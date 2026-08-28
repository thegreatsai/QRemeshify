"""Import data out of an existing site .xlsm into the new schema.

Two things are imported here, both cross-checked against the workbook's own
formulas (not guessed) before being trusted:

1. 'Data Lists' -> reference_list/reference_item. Every non-empty column
   becomes a ReferenceList keyed by its slugified header, with each
   non-blank cell below it becoming a ReferenceItem. This is a faithful,
   lossless move of the sheet's dropdown data into the database.

2. 'School Information' -> Site.building_code / Site.rack_id. These two
   cells are the only ones in the sheet that other sheets actually pull
   from (every header across the workbook reads
   'School Information'!$F$4 for building code and $F$5 for rack ID), so
   they're the only fields imported with confidence. The remaining labels
   in that sheet (Street Address, DOE FPM Name, Checkout Status, Student
   Pop., Multi-School Status) are visible in column A but never referenced
   by formula elsewhere in the workbook, so their value column could not be
   verified this way -- add them to SCHOOL_INFO_FIELDS once confirmed
   against a filled-in site file, rather than guessing coordinates.
"""

import re
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from app.models.reference_data import ReferenceItem, ReferenceList
from app.models.site import Site

DATA_LISTS_SHEET = "Data Lists"
SCHOOL_INFO_SHEET = "School Information"

# label row -> (Site field name, cell). Confirmed via cross-sheet formula
# references (see module docstring); extend only after similar verification.
SCHOOL_INFO_FIELDS: dict[str, str] = {
    "building_code": "F4",
    "rack_id": "F5",
}


def _slugify(header: str) -> str:
    key = re.sub(r"[^\w]+", "_", header.strip().lower())
    return key.strip("_")


def import_reference_lists(
    xlsm_path: str | Path, db: Session, sheet_name: str = DATA_LISTS_SHEET
) -> dict[str, int]:
    """Import every column of the Data Lists sheet as a reference list.

    Returns a dict of list key -> number of new items added. Idempotent:
    re-running against the same (or an updated) file only adds values not
    already present.
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Random cell access (ws.cell(row, col)) is pathologically slow on a
    # read-only worksheet this large -- it re-walks the XML stream per call.
    # A single sequential pass with iter_rows() is the supported fast path.
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter)
    headers: dict[int, str] = {
        col: str(value).strip() for col, value in enumerate(header_row) if value
    }

    columns: dict[int, list[str]] = {col: [] for col in headers}
    for row in rows_iter:
        for col in headers:
            if col >= len(row):
                continue
            raw = row[col]
            if raw is None or str(raw).strip() == "":
                continue
            columns[col].append(str(raw).strip())

    added_counts: dict[str, int] = {}
    for col, header in headers.items():
        key = _slugify(header)
        if not key:
            continue

        ref_list = db.query(ReferenceList).filter(ReferenceList.key == key).first()
        if ref_list is None:
            ref_list = ReferenceList(key=key, label=header)
            db.add(ref_list)
            db.flush()

        existing_values = {item.value for item in ref_list.items}
        sort_order = len(existing_values)
        added = 0
        for value in columns[col]:
            if value in existing_values:
                continue
            db.add(ReferenceItem(list=ref_list, value=value, label=value, sort_order=sort_order))
            existing_values.add(value)
            sort_order += 1
            added += 1
        added_counts[key] = added

    db.commit()
    wb.close()
    return added_counts


def import_site(xlsm_path: str | Path, db: Session, name: str, district: str | None = None) -> Site:
    """Import a single site's identity from School Information.

    `name` and `district` aren't reliably extractable from the sheet (see
    module docstring) so they're passed in by the caller -- typically from
    the filename or a manifest, not guessed from unverified cells.
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, read_only=True)
    ws = wb[SCHOOL_INFO_SHEET]

    values = {
        field: ws[cell].value for field, cell in SCHOOL_INFO_FIELDS.items()
    }
    wb.close()

    building_code = values.get("building_code")
    if not building_code:
        raise ValueError(
            f"'{SCHOOL_INFO_SHEET}'!{SCHOOL_INFO_FIELDS['building_code']} is empty in {xlsm_path} "
            "-- this looks like a blank template, not a filled-in site file"
        )

    site = db.query(Site).filter(Site.building_code == str(building_code)).first()
    if site is None:
        site = Site(building_code=str(building_code), name=name, district=district)
        db.add(site)
    site.rack_id = str(values["rack_id"]) if values.get("rack_id") else site.rack_id
    db.commit()
    db.refresh(site)
    return site
